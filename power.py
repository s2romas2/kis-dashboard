#!/usr/bin/env python3
# ⚡ 미국 전력 수요 파이프라인 수집기
# ① EIA-860M 계획 발전소(월간 공식): 연료별·가동예정연도별·주별 GW 스냅샷 (히스토리 누적)
# ② 뉴스 피드: 대형부하(데이터센터) 계통 신청 / 유틸리티 파이프라인 / 장비 백로그 / CSP 캐펙스
# 결과: public/data/power.json
import os, sys, json, time, re, io, urllib.request, urllib.parse, ssl, datetime
import xml.etree.ElementTree as ET

OUT = 'public/data/power.json'
UA = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/126'}
CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE
DEBUG = []

MONTHS = ['january', 'february', 'march', 'april', 'may', 'june',
          'july', 'august', 'september', 'october', 'november', 'december']

def fetch(url, t=90, tries=2):
    for i in range(tries):
        try:
            return urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=t, context=CTX).read()
        except Exception:
            time.sleep(2)
    return b''

def parse_860m(b):
    """엑셀 → {'total': GW, 'fuel': {tech: GW}, 'year': {yyyy: GW}, 'state': {st: GW}, 'gas': GW, 'nuke': GW, 'n': 행수}"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True)
    if 'Planned' not in wb.sheetnames:
        return None
    ws = wb['Planned']
    hdr = None
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        if r and any('Capacity' in str(c or '') for c in r):
            hdr = [str(c or '') for c in r]
            hdr_row = i + 1
            break
    if not hdr:
        return None
    def col(*names):
        for j, n in enumerate(hdr):
            if any(k in n for k in names):
                return j
        return None
    c_cap, c_tech, c_yr, c_st = col('Capacity'), col('Technology'), col('Planned Operation Year', 'Year'), col('Plant State', 'State')
    total = 0.0
    fuel, year, state = {}, {}, {}
    n = 0
    for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        try:
            cap = float(r[c_cap])
        except Exception:
            continue
        n += 1
        total += cap
        t = str(r[c_tech])[:40] if c_tech is not None and r[c_tech] else '기타'
        fuel[t] = fuel.get(t, 0) + cap
        if c_yr is not None and r[c_yr]:
            y = str(r[c_yr])[:4]
            if y.isdigit():
                year[y] = year.get(y, 0) + cap
        if c_st is not None and r[c_st]:
            s = str(r[c_st])[:2]
            state[s] = state.get(s, 0) + cap
    gw = lambda d: {k: round(v / 1000, 2) for k, v in d.items()}
    fuel = dict(sorted(gw(fuel).items(), key=lambda x: -x[1]))
    gas = sum(v for k, v in fuel.items() if 'Natural Gas' in k)
    nuke = sum(v for k, v in fuel.items() if 'Nuclear' in k)
    return {'total': round(total / 1000, 1), 'fuel': fuel,
            'year': dict(sorted(gw(year).items())),
            'state': dict(sorted(gw(state).items(), key=lambda x: -x[1])[:12]),
            'gas': round(gas, 1), 'nuke': round(nuke, 1), 'n': n}

def eia_urls():
    """최신월부터 과거로 (본 페이지 + 아카이브 경로 조합)"""
    today = datetime.date.today()
    out = []
    y, m = today.year, today.month
    for _ in range(14):
        mn = MONTHS[m - 1]
        out.append(('%04d-%02d' % (y, m), [
            'https://www.eia.gov/electricity/data/eia860m/xls/%s_generator%d.xlsx' % (mn, y),
            'https://www.eia.gov/electricity/data/eia860m/archive/xls/%s_generator%d.xlsx' % (mn, y)]))
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    return out

def collect_eia(prev_snap, backfill=False):
    snap = dict(prev_snap or {})
    fetched = 0
    for key, urls in eia_urls():
        if key in snap:
            if not backfill:
                break  # 최신이 이미 있으면 종료(과거는 이미 수집됨)
            continue
        got = None
        for u in urls:
            b = fetch(u, 120)
            if b[:2] == b'PK':
                got = b
                break
        if not got:
            continue  # 그 달 미발표(HTML 응답)
        try:
            r = parse_860m(got)
            if r and r['total'] > 50:  # 정합성: 미국 계획 발전이 50GW 미만일 수 없음
                snap[key] = r
                fetched += 1
                DEBUG.append('EIA %s: %sGW(가스 %s·원전 %s)' % (key, r['total'], r['gas'], r['nuke']))
        except Exception as e:
            DEBUG.append('EIA %s 파싱 실패 %r' % (key, str(e)[:40]))
        if not backfill and fetched:
            break
        if fetched >= 12:
            break
        time.sleep(1)
    return snap

# ── 뉴스 피드 (구글뉴스 RSS + 번역) ──
def rss(lang, query):
    if lang == 'ko':
        u = 'https://news.google.com/rss/search?q=%s&hl=ko&gl=KR&ceid=KR:ko' % urllib.parse.quote(query)
    else:
        u = 'https://news.google.com/rss/search?q=%s&hl=en-US&gl=US&ceid=US:en' % urllib.parse.quote(query)
    out = []
    try:
        root = ET.fromstring(fetch(u, 30).decode('utf-8', 'ignore'))
        for it in list(root.iter('item'))[:7]:
            t = (it.findtext('title') or '').strip()
            lk = (it.findtext('link') or '').strip()
            pd = (it.findtext('pubDate') or '')
            try:
                d = time.strftime('%Y-%m-%d', time.strptime(pd[5:16], '%d %b %Y'))
            except Exception:
                d = ''
            if t and lk:
                out.append({'d': d, 't': t[:150], 'u': lk, 'l': lang})
    except Exception:
        pass
    return out

def tr_ko(text):
    try:
        u = 'https://api.mymemory.translated.net/get?q=%s&langpair=en|ko' % urllib.parse.quote(text[:300])
        r = json.loads(fetch(u, 20, 1).decode('utf-8'))
        t = (r.get('responseData') or {}).get('translatedText') or ''
        if t and 'MYMEMORY' not in t.upper():
            return t
    except Exception:
        pass
    return None

# ── 미국 PPI (생산자물가 — 품목별) : FRED 미러(무키) 우선, BLS API 폴백 ──
PPI_SERIES = {
 'PCU335311335311': '변압기',
 'PCU335313335313': '배전반·개폐기',
 'PCU335312335312': '모터·발전기',
 'PCU333611333611': '터빈·발전기세트',
 'PCU335929335929': '전력·통신 전선',
}

def collect_ppi(prev):
    out = dict(prev or {})
    for sid, nm in PPI_SERIES.items():
        ser = dict((out.get(sid) or {}).get('d') or {})
        got = False
        # ① FRED csv (무키)
        try:
            x = fetch('https://fred.stlouisfed.org/graph/fredgraph.csv?id=%s' % sid, 45, 1).decode('utf-8', 'ignore')
            for ln in x.strip().split('\n')[1:]:
                p = ln.split(',')
                if len(p) >= 2 and p[1] not in ('.', ''):
                    try:
                        ser[p[0][:7]] = float(p[1])
                        got = True
                    except Exception:
                        pass
        except Exception:
            pass
        if got:
            out[sid] = {'n': nm, 'd': ser}
            DEBUG.append('PPI %s(FRED) %d점' % (nm, len(ser)))
        time.sleep(0.5)
    if not any((out.get(s) or {}).get('d') for s in PPI_SERIES):
        # ② BLS API 폴백 (BLS_KEY 있으면 사용)
        try:
            key = os.environ.get('BLS_KEY', '')
            body = {'seriesid': list(PPI_SERIES.keys()), 'startyear': '2020',
                    'endyear': str(datetime.date.today().year)}
            if key:
                body['registrationkey'] = key
            req = urllib.request.Request('https://api.bls.gov/publicAPI/v2/timeseries/data/',
                                         json.dumps(body).encode(),
                                         {'Content-Type': 'application/json', 'User-Agent': UA['User-Agent']})
            r = json.loads(urllib.request.urlopen(req, timeout=60, context=CTX).read())
            for srs in r.get('Results', {}).get('series', []):
                sid = srs['seriesID']
                ser = dict((out.get(sid) or {}).get('d') or {})
                for d in srs.get('data') or []:
                    if d['period'].startswith('M'):
                        ser['%s-%s' % (d['year'], d['period'][1:])] = float(d['value'])
                if ser:
                    out[sid] = {'n': PPI_SERIES[sid], 'd': ser}
            DEBUG.append('PPI BLS 폴백: %s' % r.get('status'))
        except Exception as e:
            DEBUG.append('PPI 실패 %r' % str(e)[:40])
    return out

# ── 한국 전력기기 수출 (관세청 nitemtrade — GW 활용신청 승인 후 채워짐) ──
CUSTOMS_KEY = os.environ.get('CUSTOMS_KEY', '')
PWR_HS = [('변압기(8504)', '8504'), ('배전반·차단기(8537)', '8537'), ('케이블(8544)', '8544')]

def collect_exports(prev):
    out = dict(prev or {})
    if not CUSTOMS_KEY:
        DEBUG.append('수출: 키 없음')
        return out
    endm = datetime.date.today().strftime('%Y%m')
    for name, hs in PWR_HS:
        try:
            q = ('/1220000/nitemtrade/getNitemtradeList?serviceKey=%s&strtYymm=202001&endYymm=%s&hsSgn=%s'
                 % (CUSTOMS_KEY, endm, hs))
            x = ''
            for base in ('http://apis.data.go.kr', 'https://apis.data.go.kr'):
                try:
                    x = urllib.request.urlopen(base + q, timeout=110).read().decode('utf-8', 'ignore')
                    if x:
                        break
                except Exception:
                    time.sleep(3)
            root = ET.fromstring(x)
            ser = dict(out.get(name) or {})
            for it in root.iter('item'):
                ym = (it.findtext('year') or '').strip()
                dlr = it.findtext('expDlr')
                m = re.search(r'(\d{4})\.(\d{2})', ym)
                if m and dlr and dlr.replace(',', '').strip().isdigit():
                    ser['%s-%s' % (m.group(1), m.group(2))] = round(int(dlr.replace(',', '')) / 1e6, 1)  # 백만$
            if ser:
                out[name] = ser
                DEBUG.append('수출 %s %d개월' % (name, len(ser)))
        except Exception as e:
            DEBUG.append('수출 %s 실패 %r' % (name, str(e)[:40]))
        time.sleep(1)
    return out

# ── 한국 전력기기 수출단가 $/kg (HS 6단위: 초고압변압기·배전반·GIS) ──
UP_HS = [('초고압 변압기(850423)', '850423'), ('배전반(853720)', '853720'), ('GIS·차단기(853521)', '853521')]

def collect_export_up(prev):
    """관세청 nitemtrade 6단위 — {품목: {ym: [수출액 백만$, 단가 $/kg]}}"""
    out = dict(prev or {})
    if not CUSTOMS_KEY:
        DEBUG.append('수출단가: 키 없음')
        return out
    endm = datetime.date.today().strftime('%Y%m')
    for name, hs in UP_HS:
        try:
            q = ('/1220000/nitemtrade/getNitemtradeList?serviceKey=%s&strtYymm=202001&endYymm=%s&hsSgn=%s'
                 % (CUSTOMS_KEY, endm, hs))
            x = ''
            for base in ('http://apis.data.go.kr', 'https://apis.data.go.kr'):
                try:
                    x = urllib.request.urlopen(base + q, timeout=110).read().decode('utf-8', 'ignore')
                    if x:
                        break
                except Exception:
                    time.sleep(3)
            root = ET.fromstring(x)
            ser = dict(out.get(name) or {})
            for it in root.iter('item'):
                ym = (it.findtext('year') or '').strip()
                dlr = (it.findtext('expDlr') or '').replace(',', '').strip()
                wgt = (it.findtext('expWgt') or '').replace(',', '').strip()
                m = re.search(r'(\d{4})\.(\d{2})', ym)
                if m and dlr.isdigit() and wgt.isdigit() and int(wgt) > 0:
                    ser['%s-%s' % (m.group(1), m.group(2))] = [round(int(dlr) / 1e6, 2), round(int(dlr) / int(wgt), 2)]
            if ser:
                out[name] = ser
                DEBUG.append('수출단가 %s %d개월' % (name, len(ser)))
        except Exception as e:
            DEBUG.append('수출단가 %s 실패 %r' % (name, str(e)[:40]))
        time.sleep(1)
    return out

# ── 미국 수입금액·한국 비중 (US Census 무역통계 — CENSUS_KEY 필요, 무료 즉시발급) ──
CENSUS_KEY = os.environ.get('CENSUS_KEY', '')
IMP_HS = [('초고압 변압기(850423)', '850423'), ('배전반(853720)', '853720'), ('GIS·차단기(853521)', '853521')]

def collect_us_imports(prev):
    """{품목: {'tot': {ym: 백만$}, 'kr': {ym: 백만$}}} — 한국 비중은 프론트에서 kr/tot"""
    out = dict(prev or {})
    if not CENSUS_KEY:
        DEBUG.append('미국수입: CENSUS_KEY 없음')
        return out
    for name, hs in IMP_HS:
        try:
            u = ('https://api.census.gov/data/timeseries/intltrade/imports/hs'
                 '?get=GEN_VAL_MO,CTY_CODE,CTY_NAME&I_COMMODITY=%s&time=from+2020-01&key=%s' % (hs, CENSUS_KEY))
            rows = json.loads(fetch(u, 90, 2).decode('utf-8', 'ignore'))
            hd = rows[0]
            iv, ic, inm, it = hd.index('GEN_VAL_MO'), hd.index('CTY_CODE'), hd.index('CTY_NAME'), hd.index('time')
            cur = out.get(name) or {}
            tot, kr = dict(cur.get('tot') or {}), dict(cur.get('kr') or {})
            for r in rows[1:]:
                try:
                    v = round(float(r[iv]) / 1e6, 1)
                except Exception:
                    continue
                if r[ic] == '-' or 'TOTAL FOR ALL' in (r[inm] or '').upper():
                    tot[r[it]] = v
                elif r[ic] == '5800':  # Korea, South
                    kr[r[it]] = v
            if tot:
                out[name] = {'tot': tot, 'kr': kr}
                DEBUG.append('미국수입 %s %d개월' % (name, len(tot)))
        except Exception as e:
            DEBUG.append('미국수입 %s 실패 %r' % (name, str(e)[:40]))
        time.sleep(0.6)
    return out

# ── 유틸리티·빅테크 CAPEX (SEC EDGAR XBRL, 무키 — 분기 확정치) ──
CAPEX_CO = [
 # (표시명, CIK, 그룹, 태그 우선순위) — 태그는 실측 검증됨
 ('AEP',        '0000004904', 'util', ['PaymentsForConstructionInProcess']),
 ('Duke',       '0001326160', 'util', ['PaymentsToAcquirePropertyPlantAndEquipment', 'PaymentsForConstructionInProcess']),
 ('Southern',   '0000092122', 'util', ['PaymentsToAcquirePropertyPlantAndEquipment', 'PaymentsForConstructionInProcess']),
 ('Dominion',   '0000715957', 'util', ['PaymentsForProceedsFromProductiveAssets', 'PaymentsForConstructionInProcess']),
 ('Microsoft',  '0000789019', 'tech', ['PaymentsToAcquirePropertyPlantAndEquipment']),
 ('Alphabet',   '0001652044', 'tech', ['PaymentsToAcquirePropertyPlantAndEquipment']),
 ('Amazon',     '0001018724', 'tech', ['PaymentsToAcquireProductiveAssets', 'PaymentsToAcquirePropertyPlantAndEquipment']),
 ('Meta',       '0001326801', 'tech', ['PaymentsToAcquirePropertyPlantAndEquipment']),
]

def collect_capex(prev):
    """10-Q/10-K 현금흐름표 설비투자(YTD 누적) → 분기값 차감 산출. {사명: {'g':그룹, 'd': {끝날짜: $B}}}"""
    out = dict(prev or {})
    hdr = {'User-Agent': 'kis-dashboard research ohseho57@gmail.com'}
    for name, cik, grp, tags in CAPEX_CO:
        usd = None
        for tag in tags:
            try:
                req = urllib.request.Request(
                    'https://data.sec.gov/api/xbrl/companyconcept/CIK%s/us-gaap/%s.json' % (cik, tag), headers=hdr)
                j = json.loads(urllib.request.urlopen(req, timeout=60, context=CTX).read())
                u = (j.get('units') or {}).get('USD') or []
                if u:
                    usd = u
                    break
            except Exception:
                pass
            time.sleep(0.3)
        if not usd:
            DEBUG.append('CAPEX %s 태그없음' % name)
            continue
        per = {}  # (start,end) -> val (뒤 공시가 덮어씀 = 최신 정정 반영)
        for e in usd:
            if e.get('form') not in ('10-Q', '10-K'):
                continue
            s, en, v = e.get('start'), e.get('end'), e.get('val')
            if s and en and isinstance(v, (int, float)):
                per[(s, en)] = v
        def days(a, b):
            return (datetime.date.fromisoformat(b) - datetime.date.fromisoformat(a)).days
        starts = {}
        for (s, en), v in per.items():
            starts.setdefault(s, []).append((en, v))
        q = {}
        for s, lst in starts.items():
            lst.sort()
            pv, pe = 0, None
            for en, v in lst:
                span = days(pe, en) if pe else days(s, en)
                if 60 <= span <= 130:  # 분기 구간만 인정 (연차 단독행 등 배제)
                    q[en] = round((v - pv) / 1e9, 2)
                pv, pe = v, en
        if q:
            ql = sorted(q.items())[-26:]
            # 태그가 끊긴 좀비 시계열 배제 (최신 분기가 15개월 이상 과거면 제외)
            if (datetime.date.today() - datetime.date.fromisoformat(ql[-1][0])).days > 450:
                DEBUG.append('CAPEX %s 스테일(%s) 제외' % (name, ql[-1][0]))
            else:
                out[name] = {'g': grp, 'd': dict(ql)}
                DEBUG.append('CAPEX %s %d분기' % (name, len(ql)))
        time.sleep(0.4)
    return out

NEWS_Q = {
 'queue': [('en', 'data center interconnection queue gigawatts'), ('en', 'ERCOT large load data center')],
 'utility': [('en', 'Dominion AEP data center pipeline gigawatts'), ('en', 'utility data center contracted load')],
 'equip': [('en', 'GE Vernova gas turbine orders backlog'), ('ko', '두산에너빌리티 가스터빈 OR SMR 수주'), ('ko', '변압기 수주 미국')],
 'capex': [('en', 'hyperscaler capex guidance data center'), ('en', 'Microsoft Google Meta Amazon capex increase')],
 'onsite': [('en', 'data center onsite power generation fuel cell'), ('en', 'nuclear SMR data center agreement')],
}

def collect_news(prev):
    out = dict(prev or {})
    budget = 10  # 번역 한도
    for k, queries in NEWS_Q.items():
        items = list(out.get(k) or [])
        seen = {re.sub(r'\W+', '', x['t'])[:60] for x in items}
        for lang, q in queries:
            for it in rss(lang, q):
                key = re.sub(r'\W+', '', it['t'])[:60]
                if key in seen:
                    continue
                seen.add(key)
                items.append(it)
            time.sleep(0.4)
        # 신규 영문 항목 번역
        for it in items:
            if budget <= 0:
                break
            if it.get('tk') or it.get('l') != 'en':
                continue
            if any('가' <= ch <= '힣' for ch in it['t']):
                continue
            t = tr_ko(it['t'])
            if t:
                it['tk'] = t[:160]
                budget -= 1
                time.sleep(0.4)
        items.sort(key=lambda x: x.get('d') or '')
        out[k] = items[-16:]
    return out

def main():
    prev = {}
    if os.path.exists(OUT):
        try:
            prev = json.load(open(OUT, encoding='utf-8'))
        except Exception:
            pass
    backfill = os.environ.get('BACKFILL') == '1'
    snap = collect_eia(prev.get('eia'), backfill)
    news = collect_news(prev.get('news'))
    ppi = collect_ppi(prev.get('ppi'))
    exp = collect_exports(prev.get('exp'))
    expup = collect_export_up(prev.get('expup'))
    usimp = collect_us_imports(prev.get('usimp'))
    capex = collect_capex(prev.get('capex'))
    if not snap and prev.get('eia'):
        snap = prev['eia']
    out = {'updated': time.strftime('%Y-%m-%d %H:%M'), 'eia': snap, 'news': news, 'ppi': ppi, 'exp': exp,
           'expup': expup, 'usimp': usimp, 'capex': capex, 'debug': DEBUG[-20:]}
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(out, open(OUT, 'w', encoding='utf-8'), ensure_ascii=False)
    print('완료', DEBUG, file=sys.stderr)

if __name__ == '__main__':
    main()
